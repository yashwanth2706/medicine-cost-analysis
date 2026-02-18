#!/usr/bin/env python3
"""
Tata 1mg Medicine URL Finder — Selenium Edition
Reads medicine names from 'medicine_prices' sheet, searches 1mg using
a real headless Chrome browser (required because 1mg is JS-rendered),
and writes found URLs into the existing 'medicine_url' column.

Setup:
    pip install selenium webdriver-manager pandas openpyxl

Usage:
    python 1mg_search.py medicines_final.xlsx
    python 1mg_search.py medicines_final.xlsx --delay 2 --start 50
    python 1mg_search.py medicines_final.xlsx --out results.xlsx
"""

import sys
import re
import time
import argparse
import logging
from pathlib import Path
from urllib.parse import quote_plus

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager


# ─── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("1mg")

# ─── Constants ────────────────────────────────────────────────────────────────
BASE_URL   = "https://www.1mg.com"
SEARCH_URL = "https://www.1mg.com/search/all?name={query}"

CLR_HEADER    = "1F4E79"
CLR_FOUND     = "E2EFDA"
CLR_NOT_FOUND = "FCE4D6"
CLR_ERROR     = "FFF2CC"
CLR_SKIP      = "F2F2F2"


# ─── Browser ──────────────────────────────────────────────────────────────────
def create_driver():
    """Launch a headless Chrome browser."""
    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--window-size=1920,1080")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=opts)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    log.info("Chrome headless browser started.")
    return driver


# ─── Search ───────────────────────────────────────────────────────────────────
def search_medicine(driver, medicine_name: str) -> dict:
    result = {
        "query":  medicine_name,
        "found":  False,
        "url":    None,
        "status": "not_found",
        "error":  None,
    }

    name = str(medicine_name).strip()
    if not name or name.lower() in ("nan", "none", ""):
        result["status"] = "skip"
        return result

    search_url = SEARCH_URL.format(query=quote_plus(name))

    try:
        driver.get(search_url)

        # Wait up to 10s for any product link to appear
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "a[href*='/drugs/'], a[href*='/otc/']")
                )
            )
        except Exception:
            pass  # Page may have loaded but with no results — check anyway

        url = _extract_url_from_page(driver)

        if url:
            result["found"]  = True
            result["url"]    = url
            result["status"] = "found"
            log.info("  ✓ %-30s → %s", name, url)
        else:
            log.info("  ✗ %-30s → not found", name)

    except Exception as e:
        result["status"] = "error"
        result["error"]  = str(e)[:120]
        log.warning("  ⚠ %-30s → %s", name, str(e)[:80])

    return result


def _extract_url_from_page(driver) -> str | None:
    """
    Pull the first valid product URL out of the current search results page.
    Tries multiple strategies since 1mg updates its HTML periodically.
    """
    page_source = driver.page_source

    # Strategy 1: Find <a> tags with product slugs (slug ends in -<digits>)
    anchors = driver.find_elements(By.TAG_NAME, "a")
    for a in anchors:
        href = a.get_attribute("href") or ""
        if re.search(r"1mg\.com/(drugs|otc|homeopathy|ayurveda)/[^/]+-\d+$", href):
            return href.rstrip("/")

    # Strategy 2: Regex over raw page source (catches JS-embedded links)
    matches = re.findall(
        r'https://www\.1mg\.com/(drugs|otc|homeopathy|ayurveda)/[a-z0-9\-]+-\d+',
        page_source
    )
    if matches:
        # matches are the capture groups, reconstruct full URL
        full_matches = re.findall(
            r'(https://www\.1mg\.com/(?:drugs|otc|homeopathy|ayurveda)/[a-z0-9\-]+-\d+)',
            page_source
        )
        if full_matches:
            return full_matches[0].rstrip("/")

    # Strategy 3: Relative path regex
    rel_matches = re.findall(
        r'["\']((\/(?:drugs|otc|homeopathy|ayurveda)\/[a-z0-9\-]+-\d+))["\']',
        page_source
    )
    if rel_matches:
        return BASE_URL + rel_matches[0][0].rstrip("/")

    return None


# ─── Excel helpers ────────────────────────────────────────────────────────────
def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _style_header(cell, text):
    cell.value = text
    cell.font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill  = PatternFill("solid", start_color=CLR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _border()


def _style_data(cell, value, colour, hyperlink=False):
    cell.value = str(value) if value not in (None, "", float("nan")) else ""
    cell.fill  = PatternFill("solid", start_color=colour)
    cell.border = _border()
    cell.alignment = Alignment(vertical="center")
    if hyperlink and value:
        cell.hyperlink = str(value)
        cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
    else:
        cell.font = Font(name="Arial", size=10)


# ─── Write results ────────────────────────────────────────────────────────────
def write_output(input_path, sheet_name, med_col, url_col, results, output_path):
    wb = openpyxl.load_workbook(input_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    # Map header names → column indices (row 1)
    headers = {str(cell.value).strip(): cell.column
               for cell in ws[1] if cell.value is not None}

    # Find medicine_name column
    med_col_idx = next(
        (idx for h, idx in headers.items() if h.lower() == med_col.lower()), None
    )
    if med_col_idx is None:
        raise ValueError(f"Column '{med_col}' not found. Headers: {list(headers.keys())}")

    # Find or create medicine_url column
    url_col_idx = next(
        (idx for h, idx in headers.items() if h.lower() == url_col.lower()), None
    )
    if url_col_idx is None:
        url_col_idx = ws.max_column + 1
        _style_header(ws.cell(row=1, column=url_col_idx), url_col)

    # Re-style all existing headers
    for col_idx in range(1, url_col_idx + 1):
        cell = ws.cell(row=1, column=col_idx)
        if cell.value:
            _style_header(cell, cell.value)
    ws.row_dimensions[1].height = 25

    # Build lookup: lowercase name → result
    result_map = {str(r["query"]).strip().lower(): r for r in results}

    # Write each data row
    for row_idx in range(2, ws.max_row + 1):
        raw = ws.cell(row=row_idx, column=med_col_idx).value
        name   = str(raw).strip() if raw else ""
        res    = result_map.get(name.lower(), {})
        status = res.get("status", "skip")
        url    = res.get("url") or ""

        colour = {"found": CLR_FOUND, "not_found": CLR_NOT_FOUND,
                  "error": CLR_ERROR, "skip": CLR_SKIP}.get(status, "FFFFFF")

        for col_idx in range(1, url_col_idx):
            c = ws.cell(row=row_idx, column=col_idx)
            c.fill   = PatternFill("solid", start_color=colour)
            c.font   = Font(name="Arial", size=10)
            c.border = _border()

        _style_data(ws.cell(row=row_idx, column=url_col_idx),
                    url, colour, hyperlink=bool(url))

    # Column widths
    col_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                col_widths[cell.column] = max(
                    col_widths.get(cell.column, 0),
                    min(len(str(cell.value)), 60)
                )
    for col_idx, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w + 3
    ws.column_dimensions[get_column_letter(url_col_idx)].width = 60
    ws.freeze_panes = "A2"

    # Summary sheet
    if "URL Summary" in wb.sheetnames:
        del wb["URL Summary"]
    ws2 = wb.create_sheet("URL Summary")
    total     = len([r for r in results if r.get("status") != "skip"])
    found     = sum(1 for r in results if r.get("status") == "found")
    not_found = sum(1 for r in results if r.get("status") == "not_found")
    errors    = sum(1 for r in results if r.get("status") == "error")

    for r_i, row in enumerate([
        ["Metric", "Count"],
        ["Total searched", total],
        ["Found on 1mg ✓", found],
        ["Not found ✗", not_found],
        ["Errors ⚠", errors],
        ["Success rate", f"{found/total*100:.1f}%" if total else "N/A"],
    ], 1):
        for c_i, val in enumerate(row, 1):
            cell = ws2.cell(row=r_i, column=c_i, value=val)
            if r_i == 1:
                _style_header(cell, val)
            else:
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(horizontal="center")
                cell.border = _border()
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 18

    wb.save(output_path)
    log.info("Saved → %s", output_path)
    return found, not_found, errors


# ─── CLI ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Search Tata 1mg for medicines in an Excel file (Selenium-based)."
    )
    parser.add_argument("xlsx",                            help="Input .xlsx file")
    parser.add_argument("--col",    default="medicine_name",   help="Medicine name column  (default: medicine_name)")
    parser.add_argument("--urlcol", default="medicine_url",    help="URL output column     (default: medicine_url)")
    parser.add_argument("--sheet",  default="medicine_prices", help="Sheet name            (default: medicine_prices)")
    parser.add_argument("--out",    default=None,              help="Output file           (default: <input>_with_urls.xlsx)")
    parser.add_argument("--delay",  type=float, default=2.0,   help="Delay between searches in seconds (default: 2)")
    parser.add_argument("--start",  type=int,   default=1,     help="Resume from row N    (default: 1)")
    args = parser.parse_args()

    input_path = Path(args.xlsx)
    if not input_path.exists():
        log.error("File not found: %s", input_path)
        sys.exit(1)

    output_path = args.out or str(input_path.with_name(input_path.stem + "_with_urls.xlsx"))

    # Read sheet
    df = pd.read_excel(str(input_path), sheet_name=args.sheet, dtype=str)
    df.columns = df.columns.str.strip()

    col_match = next((c for c in df.columns if c.lower() == args.col.lower()), None)
    if col_match is None:
        log.error("Column '%s' not found. Available: %s", args.col, list(df.columns))
        sys.exit(1)

    all_medicines = df[col_match].fillna("").tolist()
    to_process    = all_medicines[args.start - 1:]

    log.info("Total rows   : %d", len(all_medicines))
    log.info("Starting row : %d", args.start)
    log.info("To process   : %d", len(to_process))

    # Launch browser once (much faster than launching per query)
    driver = create_driver()

    results = []
    try:
        for i, name in enumerate(to_process, start=args.start):
            log.info("[%d/%d] %s", i, len(all_medicines), name)
            results.append(search_medicine(driver, name))
            if i < len(all_medicines):
                time.sleep(args.delay)
    finally:
        driver.quit()
        log.info("Browser closed.")

    # Pad skips for rows before --start
    if args.start > 1:
        prefix = [{"query": n, "status": "skip", "url": None, "found": False, "error": None}
                  for n in all_medicines[:args.start - 1]]
        results = prefix + results

    found, not_found, errors = write_output(
        str(input_path), args.sheet, col_match, args.urlcol, results, output_path
    )

    total = len([r for r in results if r.get("status") != "skip"])
    print("\n" + "═" * 58)
    print(f"  Total searched : {total}")
    print(f"  ✓ Found        : {found}" + (f"  ({found/total*100:.1f}%)" if total else ""))
    print(f"  ✗ Not found    : {not_found}")
    print(f"  ⚠ Errors       : {errors}")
    print(f"  Output         : {output_path}")
    print("═" * 58)


if __name__ == "__main__":
    sys.exit(main())