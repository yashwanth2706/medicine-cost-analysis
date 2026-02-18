#!/usr/bin/env python3
"""
Medicine Price Updater
Reads medicine_prices sheet from an xlsx, fetches MRP for each URL
by calling fetch_cost_by_url.py, and writes prices back to the file.

Usage:
    python update_medicine_prices.py <path_to_xlsx>
    python update_medicine_prices.py medicine_with_pricing_and_urls.xlsx

The script:
  - Reads the 'medicine_prices' sheet
  - Iterates rows while column A ('#') has a value
  - Skips rows with no URL (logs as SKIPPED)
  - Skips rows that already have a price (logs as ALREADY_FILLED)
  - Calls: python fetch_cost_by_url.py "<url>" -mrp
  - Writes returned MRP into column D (medicine_price)
  - Saves after every row so progress is not lost on crash
  - Prints a summary at the end
"""

import sys
import os
import subprocess
import time
from datetime import datetime

import openpyxl


# ── Config ────────────────────────────────────────────────────────────────────

SHEET_NAME      = "medicine_prices"
COL_NUM         = 1   # A  → '#'
COL_NAME        = 2   # B  → medicine_name
COL_URL         = 3   # C  → medicine_url
COL_PRICE       = 4   # D  → medicine_price

# Path to the fetcher script (assumed to be in same dir, or adjust as needed)
FETCHER_SCRIPT  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fetch_cost_by_url.py")

# Delay between requests (seconds) — be polite to 1mg servers
REQUEST_DELAY   = 1.5


# ── Logging helpers ───────────────────────────────────────────────────────────

def log(row_num, name, status, detail=""):
    ts = datetime.now().strftime("%H:%M:%S")
    detail_str = f"  →  {detail}" if detail else ""
    print(f"[{ts}]  Row {row_num:>4}  |  #{int(row_num)-1:<4}  |  {status:<14}  |  {name:<45}  {detail_str}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python update_medicine_prices.py <path_to_xlsx>")
        sys.exit(1)

    xlsx_path = sys.argv[1]

    if not os.path.exists(xlsx_path):
        print(f"[ERROR] File not found: {xlsx_path}")
        sys.exit(1)

    if not os.path.exists(FETCHER_SCRIPT):
        print(f"[ERROR] Fetcher script not found: {FETCHER_SCRIPT}")
        print("        Make sure fetch_cost_by_url.py is in the same folder.")
        sys.exit(1)

    print(f"\n{'='*80}")
    print(f"  Medicine Price Updater")
    print(f"  File   : {xlsx_path}")
    print(f"  Sheet  : {SHEET_NAME}")
    print(f"  Script : {FETCHER_SCRIPT}")
    print(f"  Start  : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*80}\n")

    wb = openpyxl.load_workbook(xlsx_path)

    if SHEET_NAME not in wb.sheetnames:
        print(f"[ERROR] Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[SHEET_NAME]

    # ── Counters ──
    total_rows      = 0
    fetched_ok      = 0
    fetched_fail    = 0
    skipped_no_url  = 0
    skipped_filled  = 0

    failed_rows = []   # list of (row_num, name, url, error)

    # ── Iterate rows ──
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        num_cell   = row[COL_NUM   - 1]
        name_cell  = row[COL_NAME  - 1]
        url_cell   = row[COL_URL   - 1]
        price_cell = row[COL_PRICE - 1]

        row_num = num_cell.row
        number  = num_cell.value

        # Stop if '#' column is empty — end of data
        if number is None:
            print(f"\n[INFO] Empty '#' cell at row {row_num} — end of data reached.\n")
            break

        total_rows += 1
        name = str(name_cell.value).strip() if name_cell.value else "(no name)"
        url  = str(url_cell.value).strip()  if url_cell.value  else ""

        # Skip rows with no URL
        if not url:
            skipped_no_url += 1
            log(row_num, name, "SKIP_NO_URL")
            continue

        # Skip rows already priced
        if price_cell.value not in (None, ""):
            skipped_filled += 1
            log(row_num, name, "ALREADY_FILLED", f"₹{price_cell.value}")
            continue

        # ── Call fetcher ──
        try:
            result = subprocess.run(
                [sys.executable, FETCHER_SCRIPT, url, "-mrp"],
                capture_output=True,
                text=True,
                timeout=30,
            )
            mrp_raw = result.stdout.strip()

            if mrp_raw and mrp_raw.replace(".", "", 1).isdigit():
                # Valid numeric price — write to cell
                mrp_value = float(mrp_raw)
                price_cell.value = mrp_value
                wb.save(xlsx_path)   # save after every successful write
                fetched_ok += 1
                log(row_num, name, "FETCHED_OK", f"₹{mrp_value}")
            else:
                # Script returned something unexpected
                error_msg = mrp_raw or result.stderr.strip() or "No output"
                fetched_fail += 1
                failed_rows.append((row_num, name, url, error_msg))
                log(row_num, name, "FETCH_FAILED", error_msg[:80])

        except subprocess.TimeoutExpired:
            fetched_fail += 1
            failed_rows.append((row_num, name, url, "Timeout after 30s"))
            log(row_num, name, "FETCH_TIMEOUT", url)

        except Exception as e:
            fetched_fail += 1
            failed_rows.append((row_num, name, url, str(e)))
            log(row_num, name, "FETCH_ERROR", str(e)[:80])

        # Polite delay between requests
        time.sleep(REQUEST_DELAY)

    # ── Summary ───────────────────────────────────────────────────────────────
    print(f"\n{'='*80}")
    print(f"  SUMMARY  —  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*80}")
    print(f"  Total rows processed   : {total_rows}")
    print(f"  ✅ Prices fetched OK   : {fetched_ok}")
    print(f"  ❌ Fetch failed        : {fetched_fail}")
    print(f"  ⏭  Skipped (no URL)   : {skipped_no_url}")
    print(f"  ⏭  Already filled     : {skipped_filled}")
    print(f"  💾 File saved to       : {xlsx_path}")
    print(f"{'='*80}")

    if failed_rows:
        print(f"\n  FAILED ROWS ({len(failed_rows)}):")
        print(f"  {'Row':<6} {'#':<6} {'Name':<45} {'Error'}")
        print(f"  {'-'*6} {'-'*6} {'-'*45} {'-'*30}")
        for r_row, r_name, r_url, r_err in failed_rows:
            r_num = r_row - 1   # data row number (1-indexed from data)
            print(f"  {r_row:<6} {r_num:<6} {r_name:<45} {r_err[:50]}")
        print()

    print(f"\n  Done.\n")
    return 0 if fetched_fail == 0 else 1


if __name__ == "__main__":
    sys.exit(main())