#!/usr/bin/env python3
"""
Tata 1mg Medicine Price Crawler (BeautifulSoup Version)
More robust HTML parsing with better product extraction
"""

import pandas as pd
import requests
import time
import re
from difflib import SequenceMatcher
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from urllib.parse import quote
from bs4 import BeautifulSoup

# Configuration
INPUT_FILE = 'medicine_list.xlsx'
OUTPUT_FILE = 'medicine_list_with_prices.xlsx'
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
}

GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')


def similarity_score(a, b):
    """Calculate similarity ratio between two strings"""
    return SequenceMatcher(None, a.lower().strip(), b.lower().strip()).ratio()


def clean_price(price_text):
    """Extract numeric price from text"""
    if not price_text:
        return None
    cleaned = re.sub(r'[^\d.]', '', str(price_text))
    try:
        return float(cleaned)
    except:
        return None


def search_tata1mg_bs4(medicine_name, retries=3):
    """
    Search Tata 1mg using BeautifulSoup for robust parsing
    Returns: (product_name, mrp_price, match_score)
    """
    for attempt in range(retries):
        try:
            url = f'https://www.1mg.com/search/all?name={quote(medicine_name)}'
            response = requests.get(url, headers=HEADERS, timeout=15)
            
            if response.status_code != 200:
                if response.status_code == 429:
                    wait = (attempt + 1) * 5
                    print(f"  ⚠ Rate limited, waiting {wait}s...")
                    time.sleep(wait)
                    continue
                return None, None, 0.0
            
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Strategy 1: Look for product cards/items
            products = soup.find_all(['div', 'li'], class_=re.compile(r'(product|item|card)', re.I))
            
            if not products:
                # Strategy 2: Look for any container with price
                products = soup.find_all(['div', 'article'], attrs={'data-price': True})
            
            for product in products[:3]:  # Check top 3 results
                # Extract product name
                name_elem = product.find(['h2', 'h3', 'a'], class_=re.compile(r'(name|title)', re.I))
                if not name_elem:
                    name_elem = product.find('a', attrs={'data-name': True})
                
                product_name = None
                if name_elem:
                    product_name = name_elem.get_text(strip=True) or name_elem.get('data-name')
                
                # Extract MRP price
                mrp_elem = product.find(['span', 'div'], string=re.compile(r'MRP.*₹', re.I))
                if not mrp_elem:
                    mrp_elem = product.find(['span', 'div'], class_=re.compile(r'(mrp|price)', re.I))
                
                mrp = None
                if mrp_elem:
                    mrp_text = mrp_elem.get_text(strip=True)
                    mrp = clean_price(mrp_text)
                
                # Check data attributes as fallback
                if not mrp and product.get('data-price'):
                    mrp = clean_price(product['data-price'])
                
                if not product_name and product.get('data-name'):
                    product_name = product['data-name']
                
                if product_name and mrp:
                    score = similarity_score(medicine_name, product_name)
                    return product_name, mrp, score
            
            # Strategy 3: Fallback - look for ANY price and name on page
            all_prices = soup.find_all(string=re.compile(r'₹\s*[\d,]+'))
            all_names = soup.find_all(['h1', 'h2', 'h3'])
            
            if all_prices and all_names:
                mrp = clean_price(all_prices[0])
                product_name = all_names[0].get_text(strip=True)
                score = similarity_score(medicine_name, product_name)
                return product_name, mrp, score
            
            return None, None, 0.0
        
        except requests.exceptions.Timeout:
            print(f"  ⚠ Timeout (attempt {attempt + 1}/{retries})")
            time.sleep(2)
        except Exception as e:
            print(f"  ⚠ Error: {str(e)[:50]}")
            return None, None, 0.0
    
    return None, None, 0.0


def crawl_medicines(input_file, output_file):
    """Main crawler function"""
    print(f"📂 Loading {input_file}...")
    df = pd.read_excel(input_file)
    
    df['found_product_name'] = ''
    df['mrp_price'] = None
    df['match_score'] = None
    df['match_type'] = ''
    
    total = len(df)
    print(f"🔍 Searching {total} medicines on Tata 1mg...\n")
    
    for idx, row in df.iterrows():
        medicine = row['medicine_name']
        print(f"[{idx + 1}/{total}] {medicine}")
        
        product_name, mrp, score = search_tata1mg_bs4(medicine)
        
        if product_name and mrp:
            df.at[idx, 'found_product_name'] = product_name
            df.at[idx, 'mrp_price'] = mrp
            df.at[idx, 'match_score'] = round(score, 2)
            
            if score >= 0.9:
                df.at[idx, 'match_type'] = 'exact'
                print(f"  ✓ EXACT: {product_name} - ₹{mrp}")
            elif score >= 0.6:
                df.at[idx, 'match_type'] = 'close'
                print(f"  ~ CLOSE ({score:.0%}): {product_name} - ₹{mrp}")
            else:
                df.at[idx, 'match_type'] = 'low_match'
                print(f"  ? LOW ({score:.0%}): {product_name} - ₹{mrp}")
        else:
            df.at[idx, 'match_type'] = 'not_found'
            print(f"  ✗ Not found")
        
        time.sleep(2)
        
        if (idx + 1) % 50 == 0:
            print(f"\n💾 Progress saved ({idx + 1}/{total})\n")
            df.to_excel(output_file, index=False)
    
    print(f"\n💾 Saving to {output_file}...")
    df.to_excel(output_file, index=False)
    
    print("🎨 Applying colors...")
    apply_colors(output_file, df)
    
    exact = len(df[df['match_type'] == 'exact'])
    close = len(df[df['match_type'] == 'close'])
    low = len(df[df['match_type'] == 'low_match'])
    missing = len(df[df['match_type'] == 'not_found'])
    
    print(f"\n✅ Complete!")
    print(f"  🟢 Exact: {exact}")
    print(f"  🟠 Close: {close} (verify)")
    print(f"  🟡 Low match: {low}")
    print(f"  ⚪ Not found: {missing}")


def apply_colors(file_path, df):
    """Apply Excel color coding"""
    wb = load_workbook(file_path)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    try:
        mrp_col = headers.index('mrp_price') + 1
    except ValueError:
        print("⚠ Column 'mrp_price' not found")
        return
    
    for idx, row_data in df.iterrows():
        excel_row = idx + 2
        match_type = row_data['match_type']
        
        if match_type == 'exact':
            ws.cell(row=excel_row, column=mrp_col).fill = GREEN_FILL
        elif match_type == 'close':
            ws.cell(row=excel_row, column=mrp_col).fill = ORANGE_FILL
        else:
            ws.cell(row=excel_row, column=mrp_col).fill = YELLOW_FILL
    
    wb.save(file_path)
    print("  ✓ Colors applied")


if __name__ == '__main__':
    print("="*60)
    print("Tata 1mg Crawler (BeautifulSoup)")
    print("="*60 + "\n")
    
    try:
        crawl_medicines(INPUT_FILE, OUTPUT_FILE)
    except KeyboardInterrupt:
        print("\n⚠ Stopped by user")
    except Exception as e:
        print(f"\n❌ Error: {e}")
