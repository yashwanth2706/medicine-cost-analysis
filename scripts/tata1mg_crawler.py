#!/usr/bin/env python3
"""
Tata 1mg Medicine Price Crawler
Crawls medicine prices from Tata 1mg and matches them to your medicine list.
"""

import pandas as pd
import requests
import time
import re
from difflib import SequenceMatcher
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from urllib.parse import quote

# Configuration
INPUT_FILE = 'medicine_list.xlsx'
OUTPUT_FILE = 'medicine_list_with_prices.xlsx'
SEARCH_URL = 'https://www.1mg.com/search/all'
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.5',
    'Referer': 'https://www.1mg.com/',
}

# Color fills for Excel
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Exact match
ORANGE_FILL = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')  # Close match
YELLOW_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Empty/Not found


def similarity_score(a, b):
    """Calculate similarity between two strings (0-1)"""
    return SequenceMatcher(None, a.lower().strip(), b.lower().strip()).ratio()


def clean_price(price_text):
    """Extract numeric price from text like '₹123.45' or 'MRP ₹123'"""
    if not price_text:
        return None
    # Remove everything except digits and decimal point
    cleaned = re.sub(r'[^\d.]', '', str(price_text))
    try:
        return float(cleaned)
    except:
        return None


def search_tata1mg(medicine_name, retries=3):
    """
    Search Tata 1mg for medicine and return (product_name, mrp_price, match_score)
    """
    for attempt in range(retries):
        try:
            # Construct search URL
            search_query = quote(medicine_name)
            url = f'https://www.1mg.com/search/all?name={search_query}'
            
            response = requests.get(url, headers=HEADERS, timeout=10)
            
            if response.status_code == 200:
                html = response.text
                
                # Parse search results (this is a simplified parser)
                # In production, use BeautifulSoup for robust HTML parsing
                
                # Try to find product name and MRP
                # Pattern 1: Look for product cards with data attributes
                product_pattern = r'data-name="([^"]+)".*?data-price="([^"]+)"'
                matches = re.findall(product_pattern, html, re.DOTALL)
                
                if matches:
                    product_name, price = matches[0]
                    mrp = clean_price(price)
                    score = similarity_score(medicine_name, product_name)
                    return product_name, mrp, score
                
                # Pattern 2: Look for MRP in price sections
                mrp_pattern = r'MRP\s*₹\s*([\d,\.]+)'
                price_matches = re.findall(mrp_pattern, html)
                
                if price_matches:
                    mrp = clean_price(price_matches[0])
                    # Try to find product name nearby
                    name_pattern = r'<h[1-3][^>]*>([^<]+)</h[1-3]>'
                    name_matches = re.findall(name_pattern, html)
                    product_name = name_matches[0] if name_matches else medicine_name
                    score = similarity_score(medicine_name, product_name)
                    return product_name, mrp, score
                
                # If no patterns matched, return None
                return None, None, 0.0
            
            elif response.status_code == 429:
                # Rate limited - wait longer
                print(f"  ⚠ Rate limited, waiting {(attempt + 1) * 5}s...")
                time.sleep((attempt + 1) * 5)
            else:
                print(f"  ⚠ HTTP {response.status_code}")
                return None, None, 0.0
        
        except requests.exceptions.Timeout:
            print(f"  ⚠ Timeout (attempt {attempt + 1}/{retries})")
            time.sleep(2)
        except Exception as e:
            print(f"  ⚠ Error: {e}")
            return None, None, 0.0
    
    return None, None, 0.0


def crawl_medicines(input_file, output_file):
    """
    Main crawler function
    """
    print(f"Loading {input_file}...")
    df = pd.read_excel(input_file)
    
    # Add columns for results
    df['found_product_name'] = ''
    df['mrp_price'] = None
    df['match_score'] = None
    df['match_type'] = ''  # 'exact', 'close', 'not_found'
    
    total = len(df)
    print(f"Total medicines to search: {total}\n")
    
    for idx, row in df.iterrows():
        medicine = row['medicine_name']
        print(f"[{idx + 1}/{total}] Searching: {medicine}")
        
        product_name, mrp, score = search_tata1mg(medicine)
        
        if product_name and mrp:
            df.at[idx, 'found_product_name'] = product_name
            df.at[idx, 'mrp_price'] = mrp
            df.at[idx, 'match_score'] = round(score, 2)
            
            # Determine match type
            if score >= 0.9:  # Very high similarity = exact match
                df.at[idx, 'match_type'] = 'exact'
                print(f"  ✓ EXACT: {product_name} - ₹{mrp}")
            elif score >= 0.6:  # Good similarity = close match
                df.at[idx, 'match_type'] = 'close'
                print(f"  ~ CLOSE ({score:.0%}): {product_name} - ₹{mrp}")
            else:
                df.at[idx, 'match_type'] = 'not_found'
                print(f"  ✗ Low match ({score:.0%}): {product_name} - ₹{mrp}")
        else:
            df.at[idx, 'match_type'] = 'not_found'
            print(f"  ✗ Not found")
        
        # Rate limiting - be respectful to the server
        time.sleep(2)  # 2 seconds between requests
        
        # Save progress every 50 medicines
        if (idx + 1) % 50 == 0:
            print(f"\n💾 Saving progress at {idx + 1}/{total}...\n")
            df.to_excel(output_file, index=False)
    
    # Final save
    print(f"\n💾 Saving final results to {output_file}...")
    df.to_excel(output_file, index=False)
    
    # Apply color coding
    print("🎨 Applying color coding...")
    apply_color_coding(output_file, df)
    
    # Summary
    exact = len(df[df['match_type'] == 'exact'])
    close = len(df[df['match_type'] == 'close'])
    not_found = len(df[df['match_type'] == 'not_found'])
    
    print(f"\n✅ Done!")
    print(f"  🟢 Exact matches: {exact}")
    print(f"  🟠 Close matches: {close} (verify these)")
    print(f"  ⚪ Not found: {not_found}")


def apply_color_coding(file_path, df):
    """
    Apply Excel color coding:
    - Green: Exact match (score >= 0.9)
    - Orange: Close match (0.6 <= score < 0.9)
    - Yellow: Not found or low score
    """
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Find column indices (assuming header row 1)
    headers = [cell.value for cell in ws[1]]
    price_col = headers.index('price (₹)') + 1 if 'price (₹)' in headers else None
    mrp_col = headers.index('mrp_price') + 1 if 'mrp_price' in headers else None
    match_col = headers.index('match_type') + 1 if 'match_type' in headers else None
    
    if not all([price_col, mrp_col, match_col]):
        print("⚠ Could not find required columns for color coding")
        return
    
    # Apply colors (starting from row 2, skipping header)
    for idx, row_data in df.iterrows():
        excel_row = idx + 2  # +2 because Excel is 1-indexed and we skip header
        match_type = row_data['match_type']
        
        if match_type == 'exact':
            ws.cell(row=excel_row, column=price_col).fill = GREEN_FILL
            ws.cell(row=excel_row, column=mrp_col).fill = GREEN_FILL
        elif match_type == 'close':
            ws.cell(row=excel_row, column=price_col).fill = ORANGE_FILL
            ws.cell(row=excel_row, column=mrp_col).fill = ORANGE_FILL
        else:
            ws.cell(row=excel_row, column=price_col).fill = YELLOW_FILL
            ws.cell(row=excel_row, column=mrp_col).fill = YELLOW_FILL
    
    wb.save(file_path)
    print("  ✓ Colors applied")


if __name__ == '__main__':
    print("="*60)
    print("Tata 1mg Medicine Price Crawler")
    print("="*60)
    print()
    
    try:
        crawl_medicines(INPUT_FILE, OUTPUT_FILE)
    except KeyboardInterrupt:
        print("\n\n⚠ Interrupted by user. Progress has been saved.")
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
