import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

df = pd.read_excel('medicine_list.xlsx')
meds = df['medicine_name'].dropna().tolist()

def normalize(name):
    name = str(name).strip().lower()
    name = re.sub(r'\s+', ' ', name)
    # remove space before unit
    name = re.sub(r'\s*(mg|mcg|ml|gm|iu|g)\b', r'\1', name)
    return name.strip()

# build mapping: original -> normalized
mapping = {med: normalize(med) for med in meds}

# deduplicate
unique_normalized = sorted(set(mapping.values()))

print(f"Before: {len(meds)} medicines")
print(f"After normalization: {len(unique_normalized)} medicines")
print(f"Removed: {len(meds) - len(unique_normalized)} duplicates")

# Save
wb = Workbook()
ws = wb.active
ws.title = "medicine_prices"

header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
header_fill = PatternFill('solid', start_color='1F4E79')
yellow_fill = PatternFill('solid', start_color='FFF2CC')
alt_fill = PatternFill('solid', start_color='EBF3FB')
white_fill = PatternFill('solid', start_color='FFFFFF')
thin = Side(style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, (h, w) in enumerate(zip(['#', 'medicine_name', 'price (₹)'], [5, 45, 15]), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    ws.column_dimensions[cell.column_letter].width = w

ws.row_dimensions[1].height = 20

for i, med in enumerate(unique_normalized, 1):
    fill = alt_fill if i % 2 == 0 else white_fill
    
    c1 = ws.cell(row=i+1, column=1, value=i)
    c1.font = Font(name='Arial', size=10)
    c1.fill = fill
    c1.alignment = Alignment(horizontal='center')
    c1.border = border
    
    c2 = ws.cell(row=i+1, column=2, value=med)
    c2.font = Font(name='Arial', size=10)
    c2.fill = fill
    c2.alignment = Alignment(horizontal='left')
    c2.border = border
    
    c3 = ws.cell(row=i+1, column=3, value='')
    c3.font = Font(name='Arial', size=10)
    c3.fill = yellow_fill
    c3.alignment = Alignment(horizontal='center')
    c3.border = border

ws.freeze_panes = 'A2'
wb.save('medicine_normalized.xlsx')
print("Saved!")