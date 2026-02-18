import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

df = pd.read_excel('medicine_list.xlsx')
meds = sorted(set(df['medicine_name'].dropna().tolist()))

def normalize_units(name):
    name = str(name).strip().lower()
    name = re.sub(r'\s+', ' ', name)
    name = re.sub(r'\s*(mg|mcg|ml|gm|iu|g)\b', r'\1', name)
    return name.strip()

normalized = [normalize_units(m) for m in meds]
norm_set = set(normalized)

UNITS = ['mg', 'ml', 'mcg', 'gm', 'iu']
bare = re.compile(r'^(.*\d)$')

mapping = {}
for med in normalized:
    if bare.match(med):
        matched = None
        for unit in UNITS:
            if med + unit in norm_set:
                matched = unit
                break
        mapping[med] = med + matched if matched else med
    else:
        mapping[med] = med

final_meds = sorted(set(mapping.values()))

# Save normalized list
wb = Workbook()
ws = wb.active
ws.title = "medicine_prices"

header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
header_fill = PatternFill('solid', start_color='1F4E79')
yellow_fill = PatternFill('solid', start_color='FFF2CC')
alt_fill = PatternFill('solid', start_color='EBF3FB')
white_fill = PatternFill('solid', start_color='FFFFFF')
thin = Side(style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, (h, w) in enumerate(zip(['#', 'medicine_name', 'price (₹)'], [5, 45, 15]), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font; cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    ws.column_dimensions[cell.column_letter].width = w
ws.row_dimensions[1].height = 20

for i, med in enumerate(final_meds, 1):
    fill = alt_fill if i % 2 == 0 else white_fill
    for col, val in enumerate([i, med, ''], 1):
        cell = ws.cell(row=i+1, column=col, value=val)
        cell.font = Font(name='Arial', size=10)
        cell.fill = yellow_fill if col == 3 else fill
        cell.alignment = Alignment(horizontal='center' if col != 2 else 'left')
        cell.border = border

ws.freeze_panes = 'A2'
wb.save('medicine_normalized.xlsx')
print(f"Final unique medicines: {len(final_meds)}")

# Also save the mapping
mapping_df = pd.DataFrame([
    {'original': k, 'normalized_to': v} 
    for k, v in sorted(mapping.items()) if k != v
])
mapping_df.to_excel('normalization_mapping.xlsx', index=False)
print(f"Mapping changes: {len(mapping_df)}")