import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

df = pd.read_excel('Task_-_DC.xlsx', sheet_name='1yr - Mem', header=0)

def normalize(name):
    name = str(name).strip().lower()
    name = re.sub(r'\s+', ' ', name)
    name = name.strip('`').strip('.')
    name = re.sub(r'\s*(mg|mcg|ml|gm|iu|g)\b', r'\1', name)
    return name.strip()

all_raw = set()
for col in ['start_medicines', 'Latest_medicines']:
    for val in df[col].dropna():
        for med in str(val).split(','):
            all_raw.add(normalize(med))

UNITS = ['mg', 'ml', 'mcg', 'gm', 'iu']
bare = re.compile(r'^(.*\d)$')

def full_normalize(name):
    n = normalize(name)
    if bare.match(n):
        for unit in UNITS:
            if n + unit in all_raw:
                return n + unit
    return n

def split_and_normalize(val):
    if pd.isna(val): return []
    meds = [full_normalize(m) for m in str(val).split(',')]
    return [m for m in meds if m] or []

rows = []
for _, row in df.iterrows():
    start_meds  = split_and_normalize(row['start_medicines'])
    latest_meds = split_and_normalize(row['Latest_medicines'])
    max_len = max(len(start_meds), len(latest_meds), 1)
    start_meds  += [None] * (max_len - len(start_meds))
    latest_meds += [None] * (max_len - len(latest_meds))
    for i in range(max_len):
        rows.append({
            'clientid':             row['clientid'],
            'start_medicine':       start_meds[i],
            'start_medicine_price': None,
            'latest_medicine':      latest_meds[i],
            'latest_medicine_price':None,
            'price_difference':     None,
            'expense_difference':   None,
        })

expanded = pd.DataFrame(rows).reset_index(drop=True)

all_meds = set()
for col in ['start_medicines', 'Latest_medicines']:
    for val in df[col].dropna():
        for med in str(val).split(','):
            n = full_normalize(med)
            if n: all_meds.add(n)
unique_meds = sorted(all_meds)

# preserve insertion order of clientids
seen = {}
ordered_clients = []
for i, row in expanded.iterrows():
    cid = row['clientid']
    if cid not in seen:
        seen[cid] = []
        ordered_clients.append(cid)
    seen[cid].append(i)

# ── styles ────────────────────────────────────────────────
thin      = Side(style='thin', color='BFBFBF')
border    = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_font  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
hdr_fill  = PatternFill('solid', start_color='1F4E79')
alt_fill  = PatternFill('solid', start_color='EBF3FB')
wht_fill  = PatternFill('solid', start_color='FFFFFF')
ylw_fill  = PatternFill('solid', start_color='FFF2CC')
grn_fill  = PatternFill('solid', start_color='E2EFDA')
cell_font = Font(name='Arial', size=10)

wb  = Workbook()
ws1 = wb.active
ws1.title = "medicines_expanded"

headers    = list(expanded.columns)
col_widths = [12, 35, 20, 35, 20, 16, 18]

# header
for c, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws1.cell(row=1, column=c, value=h)
    cell.font = hdr_font; cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border
    ws1.column_dimensions[get_column_letter(c)].width = w
ws1.row_dimensions[1].height = 22

# ── write all data rows first ─────────────────────────────
color_toggle = False
prev_cid = None

for df_idx, row_data in expanded.iterrows():
    r   = df_idx + 2
    cid = row_data['clientid']
    if cid != prev_cid:
        color_toggle = not color_toggle
        prev_cid = cid
    base = alt_fill if color_toggle else wht_fill

    specs = [
        (1, cid,                         base,     'left'),
        (2, row_data['start_medicine'],  base,     'left'),
        (3, None,                         ylw_fill, 'center'),
        (4, row_data['latest_medicine'], base,     'left'),
        (5, None,                         ylw_fill, 'center'),
        (6, None,                         ylw_fill, 'center'),
        (7, None,                         grn_fill, 'center'),
    ]
    for col, val, fill, align in specs:
        cell = ws1.cell(row=r, column=col, value=val)
        cell.font = cell_font; cell.fill = fill; cell.border = border
        cell.alignment = Alignment(horizontal=align, vertical='center')

# ── merge F and G per clientid AFTER all rows written ─────
for cid in ordered_clients:
    indices    = seen[cid]
    excel_rows = [i + 2 for i in indices]
    first_r    = excel_rows[0]
    last_r     = excel_rows[-1]

    if len(excel_rows) > 1:
        ws1.merge_cells(f'F{first_r}:F{last_r}')
        ws1.merge_cells(f'G{first_r}:G{last_r}')

    # top-left cell of merge — empty, styled only
    for col, fill in [(6, ylw_fill), (7, grn_fill)]:
        cell = ws1.cell(row=first_r, column=col)
        cell.value = None
        cell.font  = cell_font
        cell.fill  = fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

ws1.freeze_panes = 'A2'

# ── Sheet 2: medicine_prices ──────────────────────────────
ws2 = wb.create_sheet("medicine_prices")

for c, (h, w) in enumerate(zip(['#', 'medicine_name', 'medicine_price (₹)'], [5, 45, 20]), 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = hdr_font; cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    ws2.column_dimensions[get_column_letter(c)].width = w
ws2.row_dimensions[1].height = 22

for i, med in enumerate(unique_meds, 1):
    fill = alt_fill if i % 2 == 0 else wht_fill
    for c, val in enumerate([i, med, None], 1):
        cell = ws2.cell(row=i+1, column=c, value=val)
        cell.font = cell_font; cell.border = border
        cell.fill = ylw_fill if c == 3 else fill
        cell.alignment = Alignment(
            horizontal='center' if c != 2 else 'left', vertical='center')

ws2.freeze_panes = 'A2'

wb.save('medicines_final.xlsx')
print(f"Done — {len(expanded)} rows | {len(unique_meds)} medicines | {len(ordered_clients)} clients")